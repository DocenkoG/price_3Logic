# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                      # Для .xlsx
import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, quoted, dump_cell, currencyType, subInParentheses, openX, sheetByName
import csv



def getXlsString(sh, i, in_columns_j):
    #print(type(sh))                    #   <class 'xlrd.sheet.Sheet'>
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]-1
        if item in ('закупка','продажа','цена со скидкой','цена_') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh) == '' :       # .find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        #if item in ('закупка','продажа','цена','цена1') :
        if item.find('цена') >= 0 :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_csv2csv( cfg ):
    inFfileName  = cfg.get('basic', 'filename_in')
    outFfileName = cfg.get('basic', 'filename_out')
    inFile  = open( inFfileName,  'r', newline='')
    outFile = open( outFfileName, 'w', newline='')
    outFields = cfg.options('cols_out')
    csvReader = csv.DictReader(inFile, delimiter=';')
    csvWriter = csv.DictWriter(outFile, fieldnames=cfg.options('cols_out'))

    print(csvReader.fieldnames)
    csvWriter.writeheader()
    recOut = {}
    for recIn in csvReader:
        for outColName in outFields :
            shablon = cfg.get('cols_out',outColName)
            for key in csvReader.fieldnames:
                if shablon.find(key) >= 0 :
                    shablon = shablon.replace(key, recIn[key])
            recOut[outColName] = shablon
        csvWriter.writerow(recOut)
    log.info('Обработано '+ str(csvReader.line_num) +'строк.')
    inFile.close()
    outFile.close()




def convert2csv(cfg):
    csvFName  = cfg.get('basic','filename_out')
    priceFName= cfg.get('basic','filename_in')
    sheetName = cfg.get('basic','sheetname')
    
    log.debug('Reading file ' + priceFName )
    sheet = sheetByName(fileName = priceFName, sheetName = sheetName)
    if not sheet :
        log.error("Нет листа "+sheetName+" в файле "+ priceFName)
        return False
    log.debug("Sheet   "+sheetName)
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)
    #brands,   discount     = config_read(cfgFName, 'discount')
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFile = open( csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriter = csv.DictWriter(outFile, fieldnames=out_cols )
    csvWriter.writeheader()

    '''                                     # Блок проверки свойств для распознавания групп      XLSX                                  
    for i in range(2393, 2397):                                                         
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['группа'] )
        print(i, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color.rgb, '------', ccc.fill.fgColor.rgb)
        print('------')
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(0, 75):                                                         
        xfx = sheet.cell_xf_index(i, 0)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 0)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 0).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut  ={}
    for i in range(1, sheet.nrows) :                                     # xls
#    for i in range(1, sheet.max_row +1) :                               # xlsx
        i_last = i
        try:
            '''
            xfx = sheet.cell_xf_index(i, 0)
            xf  = book.xf_list[xfx]
            level = xf.alignment.indent_level
            bgci  = xf.background.pattern_colour_index
            ccc   = sheet.cell(i, 0)
            value = ccc.value   
            '''
#            impValues = getXlsxString(sheet, i, in_cols_j)
            impValues = getXlsString(sheet, i, in_cols_j)
            try:                                      # Пустое поле наследует значение из предыдущей строки.
                if  impValues["grp1"] =="":  impValues["grp1"] = grp1  
                else: grp1 = impValues["grp1"]
            except Exception as e:
                pass
            for outColName in out_template.keys() :
                shablon = out_template[outColName]
                for key in impValues.keys():
                    if shablon.find(key) >= 0 :
                        shablon = shablon.replace(key, impValues[key])
                if (outColName in ("закупка","продажа")) and ("*" in shablon) :
                    p = shablon.find("*")
                    vvv1 = float(shablon[:p])
                    vvv2 = float(shablon[p+1:])
                    shablon = str(round(vvv1 * vvv2, 2))
                recOut[outColName] = shablon.strip()

            if recOut['код'].lower() in ('','model',"артикул","sku") :       # Пустая строка
                #print (i, 'Пустая строка!!!')
                continue
            csvWriter.writerow(recOut)

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    outFile.close()



def download( cfgName ):
    basicNamelist, basic = config_read( cfgName, 'basic' )
    fUnitName = cfg.get('download', 'unittest')
    pathDwnld = './tmp'
    pathPython2 = 'c:/Python27/python.exe'
    retCode = False
    FnewName = 'new_auvix_dealer.csv'
    if  not os.path.exists(fUnitName):
        log.debug( 'Отсутствует юниттест для загрузки прайса ' + fUnitName)
    else:
        dir_befo_download = set(os.listdir(pathDwnld))
        os.system( pathPython2 + ' ' + fUnitName)                                                           # Вызов unittest'a
        dir_afte_download = set(os.listdir(pathDwnld))
        new_files = list( dir_afte_download.difference(dir_befo_download))
        if len(new_files) == 1 :   
            new_file = new_files[0]                                                     # загружен ровно один файл. 
            new_ext  = os.path.splitext(new_file)[-1]
            DnewFile = os.path.join( pathDwnld,new_file)
            new_file_date = os.path.getmtime(DnewFile)
            log.info( 'Скачанный файл ' +DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) ) )
            if new_ext == '.zip':                                                       # Архив. Обработка не завершена
                log.debug( 'Zip-архив. Разархивируем.')
                work_dir = os.getcwd()                                                  
                os.chdir( os.path.join( pathDwnld ))
                dir_befo_download = set(os.listdir(os.getcwd()))
                os.system('unzip -oj ' + new_file)
                os.remove(new_file)   
                dir_afte_download = set(os.listdir(os.getcwd()))
                new_files = list( dir_afte_download.difference(dir_befo_download))
                if len(new_files) == 1 :   
                    new_file = new_files[0]                                             # разархивирован ровно 2 файл. 
                    new_ext  = os.path.splitext(new_file)[-1]
                    DnewFile = os.path.join( os.getcwd(),new_file)
                    new_file_date = os.path.getmtime(DnewFile)
                    log.debug( 'Файл из архива ' +DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) )     )
                    DnewPrice = DnewFile
                elif len(new_files) >1 :
                    log.debug( 'В архиве не два файла. Надо разбираться.')
                    DnewPrice = "dummy"
                else:
                    log.debug( 'Нет новых файлов после разархивации. Загляни в папку юниттеста поставщика.')
                    DnewPrice = "dummy"
                os.chdir(work_dir)
            elif new_ext not in ( '.csv', '.htm', '.xls', '.xlsx', 'xlsb'):
                DnewPrice = DnewFile                                             # Имя скачанного прайса
            if DnewPrice != "dummy" :
                FoldName = 'old_auvix_dealer.csv'                         # Старая копия прайса, для сравнения даты
                FnewName = 'new_auvix_dealer.csv'                         # Предыдущий прайс, с которым работает макрос
                if  (not os.path.exists( FnewName)) or new_file_date > os.path.getmtime(FnewName) : 
                    log.debug( 'Предыдущего прайса нет или скачанный файл новее. Копируем его.' )
                    if os.path.exists( FoldName): os.remove( FoldName)
                    if os.path.exists( FnewName): os.rename( FnewName, FoldName)
                    shutil.copy2(DnewPrice, FnewName)
                    retCode = True
                else:
                    log.debug( 'Файл у поставщика старый, копироавать его не надо.' )
                # Убрать скачанные файлы
                if  os.path.exists(DnewPrice):  os.remove(DnewPrice)   
            
        elif len(new_files) == 0 :        
            log.debug( 'Не удалось скачать файл прайса ')
        else:
            log.debug( 'Скачалось несколько файлов. Надо разбираться ...')
    return FnewName,retCode



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    if price_datetime+qty_seconds < time.time() :
        file_age = round((time.time()-price_datetime)/24/60/60)
        log.error('Файл "'+fileName+'" устарел!  Допустимый период '+ str(qty_days)+' дней, а ему ' + str(file_age) )
        return False
    else:
        return True



def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('confidential.cfg'):     
        cfg.read('confidential.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def processing(cfgFName):
    log.info('----------------------- Processing '+cfgFName )
    cfg = config_read(cfgFName)
    csvFName  = cfg.get('basic','filename_out')
    priceFName= cfg.get('basic','filename_in')
    
    if cfg.has_section('download'):
        result = download(cfg)
    if is_file_fresh( priceFName, int(cfg.get('basic','срок годности'))):
        #os.system( dealerName + '_converter_xlsx.xlsm')
        convert2csv(cfg)



def main( dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый 
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          '+dealerName )
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            processing(cfgFName)


if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
